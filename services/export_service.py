from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


COLOR_CABECERA = "D8E3D2"
FORMATO_MONEDA = '$ #,##0.00'
FORMATO_USD = 'USD #,##0.00'
FORMATO_PORCENTAJE = '0.00"%"'


def generar_excel_presupuesto(presupuesto: dict) -> BytesIO:
    libro = Workbook()
    hoja_inicial = libro.active
    libro.remove(hoja_inicial)

    for escenario in presupuesto["escenarios"]:
        hoja = libro.create_sheet(title=_limitar_nombre_hoja(escenario["nombre"]))
        _escribir_hoja_escenario(hoja, presupuesto, escenario)

    if len(presupuesto["escenarios"]) == 2:
        hoja_comparacion = libro.create_sheet(title="Comparacion")
        _escribir_hoja_comparacion(hoja_comparacion, presupuesto)

    buffer = BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer


def generar_pdf_presupuesto(presupuesto: dict) -> BytesIO:
    buffer = BytesIO()
    documento = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=presupuesto["nombre_proyecto"],
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloMarca",
        parent=estilos["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=24,
        leading=28,
        textColor=colors.HexColor("#122620"),
        spaceAfter=8,
    )
    subtitulo = ParagraphStyle(
        "SubtituloMarca",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#365949"),
        spaceAfter=6,
    )
    texto = ParagraphStyle(
        "TextoMarca",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#23312d"),
    )

    story = [
        Paragraph("Firmamento / Orbita Studio", subtitulo),
        Paragraph("Presupuesto editorial", titulo),
        Paragraph(
            f"<b>Proyecto:</b> {presupuesto['nombre_proyecto']}<br/>"
            f"<b>Cliente:</b> {presupuesto['cliente']}<br/>"
            f"<b>Fecha:</b> {presupuesto['fecha']}",
            texto,
        ),
    ]
    if presupuesto.get("notas"):
        story.extend([Spacer(1, 4 * mm), Paragraph(f"<b>Notas:</b> {presupuesto['notas']}", texto)])

    for escenario in presupuesto["escenarios"]:
        story.extend([Spacer(1, 8 * mm), Paragraph(escenario["nombre"], estilos["Heading2"])])
        resumen = Table(
            [
                ["Subtotal", _formatear_moneda_ars(escenario["subtotal"]), "Total", _formatear_moneda_ars(escenario["total"])],
                [
                    "Precio por ejemplar",
                    _formatear_moneda_ars(escenario["precio_por_ejemplar"]),
                    "Precio USD",
                    _formatear_moneda_usd(escenario["precio_usd"]),
                ],
                ["Margen", _formatear_porcentaje(escenario["porcentaje_ganancia"]), "Ganancia neta", _formatear_moneda_ars(escenario["ganancia_neta_pesos"])],
            ],
            colWidths=[40 * mm, 40 * mm, 40 * mm, 45 * mm],
        )
        resumen.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f5f1e8")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#122620")),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d9d2c4")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#d9d2c4")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(resumen)
        story.append(Spacer(1, 4 * mm))

        tabla_items = [["Concepto", "Monto ARS", "Nota"]]
        for item in escenario["items"]:
            tabla_items.append([
                item["nombre"],
                _formatear_moneda_ars(item["monto"]),
                item.get("nota") or "-",
            ])
        tabla = Table(tabla_items, colWidths=[55 * mm, 35 * mm, 70 * mm], repeatRows=1)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d8e3d2")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#122620")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d9d2c4")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(tabla)

    documento.build(story)
    buffer.seek(0)
    return buffer


def _escribir_hoja_escenario(hoja, presupuesto: dict, escenario: dict) -> None:
    hoja["A1"] = "Proyecto"
    hoja["B1"] = presupuesto["nombre_proyecto"]
    hoja["A2"] = "Cliente"
    hoja["B2"] = presupuesto["cliente"]
    hoja["A3"] = "Fecha"
    hoja["B3"] = presupuesto["fecha"]
    hoja["A5"] = f"Escenario: {escenario['nombre']}"

    encabezados = ["Concepto", "Monto ARS", "Nota"]
    for columna, encabezado in enumerate(encabezados, start=1):
        celda = hoja.cell(row=7, column=columna, value=encabezado)
        _estilizar_cabecera(celda)

    fila = 8
    for item in escenario["items"]:
        hoja.cell(row=fila, column=1, value=item["nombre"])
        monto = hoja.cell(row=fila, column=2, value=item["monto"])
        monto.number_format = FORMATO_MONEDA
        hoja.cell(row=fila, column=3, value=item.get("nota"))
        fila += 1

    fila += 1
    resultados = [
        ("Subtotal", escenario["subtotal"]),
        ("% Ganancia", escenario["porcentaje_ganancia"]),
        ("Total", escenario["total"]),
        ("Precio por ejemplar", escenario["precio_por_ejemplar"]),
        ("Precio USD", escenario["precio_usd"]),
        ("Ganancia neta", escenario["ganancia_neta_pesos"]),
    ]
    for etiqueta, valor in resultados:
        hoja.cell(row=fila, column=1, value=etiqueta)
        celda = hoja.cell(row=fila, column=2, value=valor)
        if etiqueta == "% Ganancia":
            celda.number_format = FORMATO_PORCENTAJE
        elif etiqueta == "Precio USD":
            celda.number_format = FORMATO_USD
        else:
            celda.number_format = FORMATO_MONEDA
        fila += 1

    hoja.column_dimensions["A"].width = 28
    hoja.column_dimensions["B"].width = 18
    hoja.column_dimensions["C"].width = 48


def _escribir_hoja_comparacion(hoja, presupuesto: dict) -> None:
    hoja["A1"] = "Comparacion de escenarios"
    hoja["A3"] = "Concepto"
    _estilizar_cabecera(hoja["A3"])

    escenario_a = presupuesto["escenarios"][0]
    escenario_b = presupuesto["escenarios"][1]

    hoja["B3"] = escenario_a["nombre"]
    hoja["C3"] = escenario_b["nombre"]
    _estilizar_cabecera(hoja["B3"])
    _estilizar_cabecera(hoja["C3"])

    filas = [
        ("Cantidad copias", escenario_a["cantidad_copias"], escenario_b["cantidad_copias"], "entero"),
        ("Subtotal", escenario_a["subtotal"], escenario_b["subtotal"], "ars"),
        ("% Ganancia", escenario_a["porcentaje_ganancia"], escenario_b["porcentaje_ganancia"], "porcentaje"),
        ("Total", escenario_a["total"], escenario_b["total"], "ars"),
        ("Precio ejemplar", escenario_a["precio_por_ejemplar"], escenario_b["precio_por_ejemplar"], "ars"),
        ("Precio USD", escenario_a["precio_usd"], escenario_b["precio_usd"], "usd"),
        ("Ganancia neta", escenario_a["ganancia_neta_pesos"], escenario_b["ganancia_neta_pesos"], "ars"),
    ]

    for indice, (etiqueta, valor_a, valor_b, formato) in enumerate(filas, start=4):
        hoja.cell(row=indice, column=1, value=etiqueta)
        celda_a = hoja.cell(row=indice, column=2, value=valor_a)
        celda_b = hoja.cell(row=indice, column=3, value=valor_b)
        if formato == "ars":
            celda_a.number_format = FORMATO_MONEDA
            celda_b.number_format = FORMATO_MONEDA
        elif formato == "usd":
            celda_a.number_format = FORMATO_USD
            celda_b.number_format = FORMATO_USD
        elif formato == "porcentaje":
            celda_a.number_format = FORMATO_PORCENTAJE
            celda_b.number_format = FORMATO_PORCENTAJE

    hoja.column_dimensions["A"].width = 24
    hoja.column_dimensions["B"].width = 20
    hoja.column_dimensions["C"].width = 20


def _estilizar_cabecera(celda) -> None:
    celda.font = Font(bold=True)
    celda.fill = PatternFill(fill_type="solid", fgColor=COLOR_CABECERA)


def _limitar_nombre_hoja(nombre: str) -> str:
    return nombre[:31] or "Escenario"


def _formatear_numero(valor: float | int | None) -> str:
    numero = float(valor or 0)
    texto = f"{numero:,.2f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def _formatear_moneda_ars(valor: float | int | None) -> str:
    return f"$ {_formatear_numero(valor)}"


def _formatear_moneda_usd(valor: float | int | None) -> str:
    return f"USD {_formatear_numero(valor)}"


def _formatear_porcentaje(valor: float | int | None) -> str:
    return f"{_formatear_numero(valor)}%"
